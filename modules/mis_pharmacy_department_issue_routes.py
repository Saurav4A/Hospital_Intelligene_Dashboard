from __future__ import annotations

from datetime import date, datetime
import io
from decimal import Decimal
from secrets import token_hex

import pandas as pd
from flask import jsonify, request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules import data_fetch


def register_mis_pharmacy_department_issue_routes(
    app,
    *,
    login_required,
    allowed_units_for_session,
    clean_df_columns,
    safe_float,
    export_cache_get_bytes,
    export_cache_put_bytes,
    excel_job_get,
    excel_job_update,
    export_executor,
    local_tz,
):
    _allowed_units_for_session = allowed_units_for_session
    _clean_df_columns = clean_df_columns
    _safe_float = safe_float
    _export_cache_get_bytes = export_cache_get_bytes
    _export_cache_put_bytes = export_cache_put_bytes
    _excel_job_get = excel_job_get
    _excel_job_update = excel_job_update
    EXPORT_EXECUTOR = export_executor
    LOCAL_TZ = local_tz

    def _coerce_int(raw_value, default: int = 0) -> int:
        try:
            return int(float(raw_value))
        except Exception:
            return int(default)

    def _safe_int(raw_value, default: int = 0) -> int:
        return _coerce_int(raw_value, default)

    def _json_scalar(value):
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime().strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return bool(value)
        if isinstance(value, int):
            return int(value)
        if isinstance(value, float):
            return float(value)
        return str(value).strip()

    def _resolve_unit(raw_unit: str):
        allowed_units = _allowed_units_for_session()
        unit = (raw_unit or "").strip().upper()
        if not allowed_units:
            return None, allowed_units, ("No unit access assigned", 403)
        if not unit:
            if len(allowed_units) == 1:
                return allowed_units[0], allowed_units, None
            return None, allowed_units, ("Please select a unit", 400)
        if unit not in allowed_units:
            return None, allowed_units, (f"Unit {unit} not permitted", 403)
        return unit, allowed_units, None

    def _parse_store_id(raw_value):
        if raw_value in {None, "", "all", "ALL"}:
            return None
        store_id = _coerce_int(raw_value, 0)
        if store_id not in {2, 3}:
            return -1
        return store_id

    def _movement_mix(issue_qty: float, return_qty: float) -> str:
        has_issue = abs(issue_qty) > 0.0005
        has_return = abs(return_qty) > 0.0005
        if has_issue and has_return:
            return "Issue + Return"
        if has_return:
            return "Return Only"
        return "Issue Only"

    def _coerce_df(value) -> pd.DataFrame:
        if isinstance(value, pd.DataFrame):
            return value
        if value is None:
            return pd.DataFrame()
        try:
            return pd.DataFrame(value)
        except Exception:
            return pd.DataFrame()

    def _prepare_rows_df(df: pd.DataFrame) -> pd.DataFrame:
        work = _clean_df_columns(_coerce_df(df)).copy()
        defaults = {
            "DepartmentID": 0,
            "DepartmentName": "Unknown Department",
            "IssuedToStoreID": 0,
            "IssuedToStoreName": "",
            "IssuedFromStoreID": 0,
            "IssuedFromStoreName": "",
            "ItemID": 0,
            "ItemName": "",
            "BatchID": 0,
            "BatchName": "",
            "Rate": 0,
            "MRP": 0,
            "IssuedQty": 0,
            "ReturnedQty": 0,
            "NetIssuedQty": 0,
            "IssuedValueRate": 0,
            "ReturnedValueRate": 0,
            "NetValueRate": 0,
            "LastMovementDate": None,
            "IssueNumber": "",
        }
        for col, default in defaults.items():
            if col not in work.columns:
                work[col] = default

        for col in ["Rate", "MRP", "IssuedQty", "ReturnedQty", "NetIssuedQty", "IssuedValueRate", "ReturnedValueRate", "NetValueRate"]:
            work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)

        work["DepartmentName"] = work["DepartmentName"].fillna("Unknown Department").astype(str).str.strip().replace("", "Unknown Department")
        work["IssuedToStoreName"] = work["IssuedToStoreName"].fillna("").astype(str).str.strip()
        work["IssuedFromStoreName"] = work["IssuedFromStoreName"].fillna("").astype(str).str.strip()
        work["ItemName"] = work["ItemName"].fillna("").astype(str).str.strip()
        work["BatchName"] = work["BatchName"].fillna("").astype(str).str.strip()
        work["IssueNumber"] = work["IssueNumber"].fillna("").astype(str).str.strip()
        work["MovementMix"] = work.apply(
            lambda row: _movement_mix(_safe_float(row.get("IssuedQty")), _safe_float(row.get("ReturnedQty"))),
            axis=1,
        )
        if "LastMovementDate" in work.columns:
            work["LastMovementDate"] = pd.to_datetime(work["LastMovementDate"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
        return work

    def _build_kpis(df: pd.DataFrame) -> dict:
        work = _prepare_rows_df(df)
        if work.empty:
            return {
                "departments": 0,
                "rows": 0,
                "issued_qty": 0.0,
                "returned_qty": 0.0,
                "net_qty": 0.0,
                "net_value_rate": 0.0,
                "opd_issued_qty": 0.0,
                "opd_returned_qty": 0.0,
                "opd_net_qty": 0.0,
                "opd_net_value_rate": 0.0,
                "ipd_issued_qty": 0.0,
                "ipd_returned_qty": 0.0,
                "ipd_net_qty": 0.0,
                "ipd_net_value_rate": 0.0,
            }

        def _store_metrics(target_store: int) -> dict:
            part = work[work["IssuedFromStoreID"].astype(int) == int(target_store)].copy()
            return {
                "issued_qty": float(part["IssuedQty"].sum()) if not part.empty else 0.0,
                "returned_qty": float(part["ReturnedQty"].sum()) if not part.empty else 0.0,
                "net_qty": float(part["NetIssuedQty"].sum()) if not part.empty else 0.0,
                "net_value_rate": float(part["NetValueRate"].sum()) if not part.empty else 0.0,
            }

        opd = _store_metrics(2)
        ipd = _store_metrics(3)
        return {
            "departments": int(work["DepartmentName"].astype(str).replace("", pd.NA).dropna().nunique()),
            "rows": int(len(work)),
            "issued_qty": float(work["IssuedQty"].sum()),
            "returned_qty": float(work["ReturnedQty"].sum()),
            "net_qty": float(work["NetIssuedQty"].sum()),
            "net_value_rate": float(work["NetValueRate"].sum()),
            "opd_issued_qty": opd["issued_qty"],
            "opd_returned_qty": opd["returned_qty"],
            "opd_net_qty": opd["net_qty"],
            "opd_net_value_rate": opd["net_value_rate"],
            "ipd_issued_qty": ipd["issued_qty"],
            "ipd_returned_qty": ipd["returned_qty"],
            "ipd_net_qty": ipd["net_qty"],
            "ipd_net_value_rate": ipd["net_value_rate"],
        }

    def _row_payload(row: dict) -> dict:
        issued_qty = _safe_float(row.get("IssuedQty"))
        returned_qty = _safe_float(row.get("ReturnedQty"))
        return {
            "department_id": _safe_int(row.get("DepartmentID"), 0),
            "department_name": _json_scalar(row.get("DepartmentName")) or "Unknown Department",
            "issued_to_store_id": _safe_int(row.get("IssuedToStoreID"), 0),
            "issued_to_store_name": _json_scalar(row.get("IssuedToStoreName")) or "",
            "issued_from_store_id": _safe_int(row.get("IssuedFromStoreID"), 0),
            "issued_from_store_name": _json_scalar(row.get("IssuedFromStoreName")) or "",
            "item_id": _safe_int(row.get("ItemID"), 0),
            "item_name": _json_scalar(row.get("ItemName")) or "",
            "batch_id": _safe_int(row.get("BatchID"), 0),
            "batch_name": _json_scalar(row.get("BatchName")) or "",
            "rate": _safe_float(row.get("Rate")),
            "mrp": _safe_float(row.get("MRP")),
            "issued_qty": issued_qty,
            "returned_qty": returned_qty,
            "net_issued_qty": _safe_float(row.get("NetIssuedQty")),
            "issued_value_rate": _safe_float(row.get("IssuedValueRate")),
            "returned_value_rate": _safe_float(row.get("ReturnedValueRate")),
            "net_value_rate": _safe_float(row.get("NetValueRate")),
            "movement_mix": _json_scalar(row.get("MovementMix")) or _movement_mix(issued_qty, returned_qty),
            "last_movement_date": _json_scalar(row.get("LastMovementDate")),
            "issue_number": _json_scalar(row.get("IssueNumber")) or "",
        }

    def _store_scope_label(store_id: int | None) -> str:
        if store_id == 2:
            return "OPD Pharmacy (2)"
        if store_id == 3:
            return "IPD Pharmacy (3)"
        return "OPD Pharmacy (2) + IPD Pharmacy (3)"

    def _store_options_payload() -> list[dict]:
        return [
            {"id": "all", "label": "All Stores"},
            {"id": 2, "label": "OPD Pharmacy (2)"},
            {"id": 3, "label": "IPD Pharmacy (3)"},
        ]

    def _build_departmental_issue_excel(
        unit: str,
        from_date: str,
        to_date: str,
        store_id: int | None,
        exported_by: str,
    ):
        exported_by = (exported_by or "Unknown").strip() or "Unknown"
        df = data_fetch.fetch_pharmacy_departmental_issue_rows(unit, from_date, to_date, store_id=store_id)
        if df is None:
            return None, None, "Department attribution path is not available for this unit/schema"
        if df.empty:
            return None, None, "No data available to export"

        work = _prepare_rows_df(df)
        kpis = _build_kpis(work)
        destination_summary_df = (
            work.assign(
                DepartmentName=work["DepartmentName"].fillna("Unknown Department").astype(str).str.strip().replace("", "Unknown Department"),
                IssuedToStoreName=work["IssuedToStoreName"].fillna("").astype(str).str.strip(),
            )
            .groupby(["DepartmentName", "IssuedToStoreName"], dropna=False, as_index=False)[["IssuedQty", "IssuedValueRate"]]
            .sum()
            .sort_values(["IssuedValueRate", "IssuedQty", "DepartmentName", "IssuedToStoreName"], ascending=[False, False, True, True])
        )
        destination_summary_df["IssuedToStoreName"] = destination_summary_df["IssuedToStoreName"].replace("", "Direct Department Issue")
        detail_df = work[[
            "DepartmentName",
            "IssuedToStoreName",
            "IssuedFromStoreName",
            "ItemName",
            "BatchName",
            "Rate",
            "MRP",
            "IssuedQty",
            "ReturnedQty",
            "NetIssuedQty",
            "IssuedValueRate",
            "ReturnedValueRate",
            "NetValueRate",
            "MovementMix",
            "LastMovementDate",
            "IssueNumber",
        ]].rename(columns={
            "DepartmentName": "Issued To Department/Store",
            "IssuedToStoreName": "Issued To Store",
            "IssuedFromStoreName": "Issued From Store",
            "ItemName": "Medicine Name",
            "BatchName": "Batch",
            "Rate": "Rate",
            "MRP": "MRP",
            "IssuedQty": "Issued Qty",
            "ReturnedQty": "Returned Qty",
            "NetIssuedQty": "Net Issued Qty",
            "IssuedValueRate": "Issued Value (Rate)",
            "ReturnedValueRate": "Returned Value (Rate)",
            "NetValueRate": "Net Value (Rate)",
            "MovementMix": "Movement Mix",
            "LastMovementDate": "Last Movement",
            "IssueNumber": "Issue Number",
        })

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        details_ws = wb.create_sheet("Details")

        title_font = Font(size=16, bold=True, color="FFFFFF")
        subtitle_font = Font(size=11, bold=False, color="E2E8F0")
        meta_label_font = Font(size=10, bold=True, color="1E3A8A")
        meta_value_font = Font(size=10, color="0F172A")
        section_font = Font(size=11, bold=True, color="1E3A8A")
        kpi_label_font = Font(size=9, bold=True, color="64748B")
        kpi_value_font = Font(size=16, bold=True, color="0F172A")
        kpi_sub_font = Font(size=9, color="475569")
        header_font = Font(size=10, bold=True, color="FFFFFF")
        white_bold_font = Font(size=10, bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="D6DFEA")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        dark_fill = PatternFill(start_color="163B73", end_color="163B73", fill_type="solid")
        header_fill = PatternFill(start_color="2D5BBA", end_color="2D5BBA", fill_type="solid")
        section_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        issue_fill = PatternFill(start_color="E9F2FF", end_color="E9F2FF", fill_type="solid")
        return_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        net_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        subtle_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        meta_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")
        borderless_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        total_cols = len(detail_df.columns)
        exported_at = datetime.now(tz=LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S")
        store_scope = _store_scope_label(store_id)

        def _apply_block_style(start_row, end_row, start_col, end_col, *, fill=None, border_obj=border):
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    cell = ws.cell(r, c)
                    if fill is not None:
                        cell.fill = fill
                    if border_obj is not None:
                        cell.border = border_obj

        def _merge_and_write(cell_range, value, *, font=None, fill=None, alignment=None, border_range=None):
            ws.merge_cells(cell_range)
            top_left = ws[cell_range.split(":")[0]]
            top_left.value = value
            if font is not None:
                top_left.font = font
            if fill is not None:
                top_left.fill = fill
            if alignment is not None:
                top_left.alignment = alignment
            if border_range is not None:
                _apply_block_style(*border_range, fill=fill)

        def _write_meta_panel(start_col, title, rows):
            end_col = start_col + 3
            _merge_and_write(
                f"{get_column_letter(start_col)}4:{get_column_letter(end_col)}4",
                title,
                font=white_bold_font,
                fill=header_fill,
                alignment=Alignment(horizontal="left", vertical="center"),
                border_range=(4, 4, start_col, end_col),
            )
            for offset, (label, value) in enumerate(rows, start=5):
                label_cell = ws.cell(offset, start_col, label)
                label_cell.font = meta_label_font
                label_cell.fill = meta_fill
                label_cell.border = border
                label_cell.alignment = Alignment(horizontal="left", vertical="center")
                _merge_and_write(
                    f"{get_column_letter(start_col + 1)}{offset}:{get_column_letter(end_col)}{offset}",
                    value,
                    font=meta_value_font,
                    fill=subtle_fill,
                    alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                    border_range=(offset, offset, start_col + 1, end_col),
                )
                ws.row_dimensions[offset].height = 22

        def _write_kpi_card(start_row, start_col, label, value, subtext, value_kind, value_fill):
            end_col = start_col + 1
            _merge_and_write(
                f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}",
                label,
                font=kpi_label_font,
                fill=subtle_fill,
                alignment=Alignment(horizontal="center", vertical="center"),
                border_range=(start_row, start_row, start_col, end_col),
            )
            _merge_and_write(
                f"{get_column_letter(start_col)}{start_row + 1}:{get_column_letter(end_col)}{start_row + 2}",
                value,
                font=kpi_value_font,
                fill=value_fill,
                alignment=Alignment(horizontal="center", vertical="center"),
                border_range=(start_row + 1, start_row + 2, start_col, end_col),
            )
            value_cell = ws.cell(start_row + 1, start_col)
            if value_kind == "money":
                value_cell.number_format = "#,##,##0.00"
            elif value_kind == "qty":
                value_cell.number_format = "#,##,##0.000"
            else:
                value_cell.number_format = "#,##,##0"
            _merge_and_write(
                f"{get_column_letter(start_col)}{start_row + 3}:{get_column_letter(end_col)}{start_row + 3}",
                subtext,
                font=kpi_sub_font,
                fill=subtle_fill,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                border_range=(start_row + 3, start_row + 3, start_col, end_col),
            )

        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "A1"

        _merge_and_write(
            "A1:H1",
            "Departmental Issue Register",
            font=title_font,
            fill=title_fill,
            alignment=Alignment(horizontal="center", vertical="center"),
            border_range=(1, 1, 1, 8),
        )
        _merge_and_write(
            "A2:H2",
            "Professional issue register export for inter-store / departmental medicine movement",
            font=subtitle_font,
            fill=dark_fill,
            alignment=Alignment(horizontal="center", vertical="center"),
            border_range=(2, 2, 1, 8),
        )
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 10

        _write_meta_panel(1, "Report Scope", [
            ("Unit", unit),
            ("Store Scope", store_scope),
            ("Date Range", f"{from_date} to {to_date}"),
        ])
        _write_meta_panel(5, "Export Details", [
            ("Exported By", exported_by),
            ("Exported At", exported_at),
            ("Copyright", "(c) ASARFI HOSPITAL"),
        ])

        _merge_and_write(
            "A9:H9",
            "Key Highlights",
            font=section_font,
            fill=section_fill,
            alignment=Alignment(horizontal="left", vertical="center"),
            border_range=(9, 9, 1, 8),
        )
        ws.row_dimensions[9].height = 22

        primary_cards = [
            ("Departments", kpis["departments"], "Unique receiving departments / stores", "count", subtle_fill),
            ("Rows", kpis["rows"], "Detailed issue movement lines", "count", subtle_fill),
            ("Issued Qty", kpis["issued_qty"], "Total quantity issued", "qty", issue_fill),
            ("Net Value", kpis["net_value_rate"], "Total issue value at rate", "money", net_fill),
        ]
        secondary_cards = [
            ("OPD Net Qty", kpis["opd_net_qty"], "Issued from OPD Pharmacy (2)", "qty", issue_fill),
            ("IPD Net Qty", kpis["ipd_net_qty"], "Issued from IPD Pharmacy (3)", "qty", issue_fill),
            ("OPD Net Value", kpis["opd_net_value_rate"], "Value from OPD Pharmacy (2)", "money", net_fill),
            ("IPD Net Value", kpis["ipd_net_value_rate"], "Value from IPD Pharmacy (3)", "money", net_fill),
        ]
        for idx, card in enumerate(primary_cards):
            _write_kpi_card(10, 1 + (idx * 2), *card)
        for idx, card in enumerate(secondary_cards):
            _write_kpi_card(15, 1 + (idx * 2), *card)
        for row in [10, 11, 12, 13, 15, 16, 17, 18]:
            ws.row_dimensions[row].height = 20

        _merge_and_write(
            "A20:H20",
            "Department / Store Wise Issue Summary",
            font=section_font,
            fill=section_fill,
            alignment=Alignment(horizontal="left", vertical="center"),
            border_range=(20, 20, 1, 8),
        )
        ws.row_dimensions[20].height = 22

        summary_header_ranges = [
            ("A21:C21", "Issued To Department/Store"),
            ("D21:E21", "Issued To Store"),
            ("F21:F21", "Issued Qty"),
            ("G21:H21", "Issued Value (Rate)"),
        ]
        for cell_range, label in summary_header_ranges:
            _merge_and_write(
                cell_range,
                label,
                font=white_bold_font,
                fill=header_fill,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                border_range=(21, 21, ws[cell_range.split(":")[0]].column, ws[cell_range.split(":")[1]].column),
            )

        table_row = 22
        for idx, row in enumerate(destination_summary_df.itertuples(index=False), start=0):
            row_fill = subtle_fill if idx % 2 == 0 else meta_fill
            _merge_and_write(
                f"A{table_row}:C{table_row}",
                getattr(row, "DepartmentName"),
                font=meta_value_font,
                fill=row_fill,
                alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                border_range=(table_row, table_row, 1, 3),
            )
            _merge_and_write(
                f"D{table_row}:E{table_row}",
                getattr(row, "IssuedToStoreName"),
                font=meta_value_font,
                fill=row_fill,
                alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                border_range=(table_row, table_row, 4, 5),
            )
            qty_cell = ws.cell(table_row, 6, getattr(row, "IssuedQty"))
            qty_cell.font = meta_value_font
            qty_cell.fill = row_fill
            qty_cell.border = border
            qty_cell.alignment = Alignment(horizontal="right", vertical="center")
            qty_cell.number_format = "#,##,##0.000"
            _merge_and_write(
                f"G{table_row}:H{table_row}",
                getattr(row, "IssuedValueRate"),
                font=meta_value_font,
                fill=row_fill,
                alignment=Alignment(horizontal="right", vertical="center"),
                border_range=(table_row, table_row, 7, 8),
            )
            ws.cell(table_row, 7).number_format = "#,##,##0.00"
            ws.row_dimensions[table_row].height = 22
            table_row += 1

        _merge_and_write(
            f"A{table_row}:E{table_row}",
            "Grand Total",
            font=Font(size=10, bold=True, color="0F172A"),
            fill=section_fill,
            alignment=Alignment(horizontal="right", vertical="center"),
            border_range=(table_row, table_row, 1, 5),
        )
        total_qty_cell = ws.cell(table_row, 6, float(destination_summary_df["IssuedQty"].sum()))
        total_qty_cell.font = Font(size=10, bold=True, color="0F172A")
        total_qty_cell.fill = section_fill
        total_qty_cell.border = border
        total_qty_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_qty_cell.number_format = "#,##,##0.000"
        _merge_and_write(
            f"G{table_row}:H{table_row}",
            float(destination_summary_df["IssuedValueRate"].sum()),
            font=Font(size=10, bold=True, color="0F172A"),
            fill=section_fill,
            alignment=Alignment(horizontal="right", vertical="center"),
            border_range=(table_row, table_row, 7, 8),
        )
        ws.cell(table_row, 7).number_format = "#,##,##0.00"
        ws.row_dimensions[table_row].height = 22

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16

        details_title_last_col = get_column_letter(total_cols)
        details_ws.merge_cells(f"A1:{details_title_last_col}1")
        details_ws["A1"] = "Departmental Issue Register - Detail Rows"
        details_ws["A1"].font = title_font
        details_ws["A1"].fill = title_fill
        details_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        details_ws.merge_cells(f"A2:{details_title_last_col}2")
        details_ws["A2"] = f"Unit: {unit} | Store Scope: {store_scope} | Date Range: {from_date} to {to_date}"
        details_ws["A2"].font = subtitle_font
        details_ws["A2"].fill = dark_fill
        details_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        details_ws.merge_cells(f"A3:{details_title_last_col}3")
        details_ws["A3"] = f"Exported By: {exported_by} | Exported At: {exported_at}"
        details_ws["A3"].font = meta_value_font
        details_ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        table_start = 5
        for col_idx, col_name in enumerate(detail_df.columns, start=1):
            cell = details_ws.cell(row=table_start, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(detail_df.itertuples(index=False), start=table_start + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = details_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                header = str(detail_df.columns[col_idx - 1])
                if header in {"Rate", "MRP", "Issued Value (Rate)", "Returned Value (Rate)", "Net Value (Rate)"}:
                    cell.number_format = "#,##,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif header in {"Issued Qty", "Returned Qty", "Net Issued Qty"}:
                    cell.number_format = "#,##,##0.000"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if header in {"Issued Qty", "Issued Value (Rate)"}:
                    cell.fill = issue_fill
                elif header in {"Returned Qty", "Returned Value (Rate)"}:
                    cell.fill = return_fill
                elif header in {"Net Issued Qty", "Net Value (Rate)"}:
                    cell.fill = net_fill
                elif row_idx % 2 == 0:
                    cell.fill = subtle_fill

        details_ws.freeze_panes = "A6"
        details_ws.auto_filter.ref = f"A{table_start}:{details_title_last_col}{details_ws.max_row}"

        detail_widths = {
            "A": 28,
            "B": 22,
            "C": 22,
            "D": 36,
            "E": 16,
            "F": 12,
            "G": 12,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 18,
            "L": 18,
            "M": 18,
            "N": 18,
            "O": 20,
            "P": 18,
        }
        for letter, width in detail_widths.items():
            details_ws.column_dimensions[letter].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"Departmental_Issue_{unit}_{from_date}_to_{to_date}.xlsx"
        return buffer.getvalue(), filename, None

    def _run_departmental_issue_export_job(
        job_id: str,
        unit: str,
        from_date: str,
        to_date: str,
        store_id: int | None,
        exported_by: str,
    ):
        _excel_job_update(job_id, state="running")
        try:
            data, filename, err = _build_departmental_issue_excel(
                unit=unit,
                from_date=from_date,
                to_date=to_date,
                store_id=store_id,
                exported_by=exported_by,
            )
            if not data:
                _excel_job_update(job_id, state="error", error=err or "No data available to export")
                return
            _export_cache_put_bytes("pharmacy_departmental_issue_xlsx_job", data, job_id)
            _excel_job_update(job_id, state="done", filename=filename)
        except Exception as exc:
            _excel_job_update(job_id, state="error", error=str(exc))

    @app.route('/api/mis/pharmacy/departmental_issue')
    @login_required(allowed_roles={"IT", "Management", "Departmental Head"})
    def api_mis_pharmacy_departmental_issue():
        unit, _, unit_error = _resolve_unit(request.args.get("unit"))
        if unit_error:
            message, code = unit_error
            return jsonify({"status": "error", "message": message}), code

        from_date = (request.args.get("from") or "").strip()
        to_date = (request.args.get("to") or "").strip()
        raw_store_id = request.args.get("store_id") or request.args.get("store")
        store_id = _parse_store_id(raw_store_id)

        if not from_date or not to_date:
            return jsonify({"status": "error", "message": "Valid from/to dates are required"}), 400
        if store_id == -1:
            return jsonify({"status": "error", "message": "Store scope must be OPD Pharmacy (2) or IPD Pharmacy (3)"}), 400

        df = data_fetch.fetch_pharmacy_departmental_issue_rows(unit, from_date, to_date, store_id=store_id)
        if df is None:
            return jsonify({"status": "error", "message": "Department attribution path is not available for this unit/schema"}), 500

        work = _prepare_rows_df(df)
        rows = [_row_payload(row) for row in work.to_dict(orient="records")] if not work.empty else []
        return jsonify({
            "status": "success",
            "unit": unit,
            "rows": rows,
            "count": len(rows),
            "kpis": _build_kpis(work),
            "store_options": _store_options_payload(),
            "filters_applied": {
                "unit": unit,
                "from": from_date,
                "to": to_date,
                "store_id": store_id,
                "store_scope": _store_scope_label(store_id),
            },
            "generated_at": datetime.now(tz=LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        })

    @app.route('/api/mis/pharmacy/departmental_issue/export_excel')
    @login_required(allowed_roles={"IT", "Management", "Departmental Head"})
    def api_mis_pharmacy_departmental_issue_export_excel():
        unit, _, unit_error = _resolve_unit(request.args.get("unit"))
        if unit_error:
            message, code = unit_error
            return jsonify({"status": "error", "message": message}), code

        from_date = (request.args.get("from") or "").strip()
        to_date = (request.args.get("to") or "").strip()
        store_id = _parse_store_id(request.args.get("store_id") or request.args.get("store"))
        if not from_date or not to_date:
            return jsonify({"status": "error", "message": "Valid from/to dates are required"}), 400
        if store_id == -1:
            return jsonify({"status": "error", "message": "Store scope must be OPD Pharmacy (2) or IPD Pharmacy (3)"}), 400

        exported_by = session.get("username") or session.get("user") or "Unknown"
        data, filename, err = _build_departmental_issue_excel(unit, from_date, to_date, store_id, exported_by)
        if not data:
            return err or "No data available to export", 404
        return send_file(
            io.BytesIO(data),
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route('/api/mis/pharmacy/departmental_issue/export_excel_job', methods=["POST"])
    @login_required(allowed_roles={"IT", "Management", "Departmental Head"})
    def api_mis_pharmacy_departmental_issue_export_excel_job():
        payload = request.get_json(silent=True) or {}
        unit, _, unit_error = _resolve_unit(payload.get("unit") or request.args.get("unit"))
        if unit_error:
            message, code = unit_error
            return jsonify({"status": "error", "message": message}), code

        from_date = (payload.get("from") or request.args.get("from") or "").strip()
        to_date = (payload.get("to") or request.args.get("to") or "").strip()
        store_id = _parse_store_id(payload.get("store_id") or request.args.get("store_id") or request.args.get("store"))
        if not from_date or not to_date:
            return jsonify({"status": "error", "message": "Valid from/to dates are required"}), 400
        if store_id == -1:
            return jsonify({"status": "error", "message": "Store scope must be OPD Pharmacy (2) or IPD Pharmacy (3)"}), 400

        exported_by = session.get("username") or session.get("user") or "Unknown"
        job_id = token_hex(16)
        _excel_job_update(job_id, state="queued", filename=None)
        EXPORT_EXECUTOR.submit(
            _run_departmental_issue_export_job,
            job_id,
            unit,
            from_date,
            to_date,
            store_id,
            exported_by,
        )
        return jsonify({"status": "queued", "job_id": job_id})

    @app.route('/api/mis/pharmacy/departmental_issue/export_excel_job_status')
    @login_required(allowed_roles={"IT", "Management", "Departmental Head"})
    def api_mis_pharmacy_departmental_issue_export_excel_job_status():
        job_id = (request.args.get("job_id") or "").strip()
        if not job_id:
            return jsonify({"status": "error", "message": "Missing job id"}), 400
        entry = _excel_job_get(job_id)
        if not entry:
            return jsonify({"status": "error", "message": "Job not found"}), 404
        return jsonify({
            "status": "success",
            "state": entry.get("state"),
            "error": entry.get("error"),
            "filename": entry.get("filename"),
        })

    @app.route('/api/mis/pharmacy/departmental_issue/export_excel_job_result')
    @login_required(allowed_roles={"IT", "Management", "Departmental Head"})
    def api_mis_pharmacy_departmental_issue_export_excel_job_result():
        job_id = (request.args.get("job_id") or "").strip()
        if not job_id:
            return "Missing job id", 400
        entry = _excel_job_get(job_id)
        if not entry:
            return "Job not found", 404
        if entry.get("state") != "done":
            return "Job not ready", 409
        data = _export_cache_get_bytes("pharmacy_departmental_issue_xlsx_job", job_id)
        if not data:
            return "Export expired", 404
        filename = entry.get("filename") or "Departmental_Issue.xlsx"
        return send_file(
            io.BytesIO(data),
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
